import base64
import os

import feapder
import json
import execjs
from openpyxl import Workbook


class AirSpiderDemo(feapder.AirSpider):
    top_file = '健康保险'
    with open('hello.js', 'r', encoding='utf-8') as f:
        js = f.read()
        ctx = execjs.compile(js)

    def start_requests(self):
        url = "https://tiaokuan.iachina.cn/sinopipi/prodtermsinfo/selectConsumerList"
        data = {
            "tagdata": "03",
            "insperdtype": "",
            "processnode": "",
            "instype": "",
            "prodpaytype": "",
            "cplbone": "PubNewProdTypeCode_02",
            "cplbtow": "",
            "cplbthree": "",
            "salestatus": "01",
            "cplbfour": "",
            "sjlxone": "",
            "sjlxtow": "",
            "sjlxthree": "",
            "proddesicode": "",
            "prodname": "",
            "inscomname": "",
            "insitemcode": "",
            "filltype": "",
            "specialattri": "",
            "insperd": "",
            "prodtypecode": "PubNewProdTypeCode_02",
            "pageNum": 1,
            "pageSize": 40
        }
        data = json.dumps(data, separators=(',', ':'))
        yield feapder.Request(url, data=data, method="POST")

    def download_midware(self, request):
        request.headers = {
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "Authorization": self.ctx.call('indentify'),
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "Content-Type": "application/json;charset=UTF-8",
            "Origin": "https://tiaokuan.iachina.cn",
            "Pragma": "no-cache",
            "Referer": "https://tiaokuan.iachina.cn/",
            "Sec-Fetch-Dest": "empty",
            "Sec-Fetch-Mode": "cors",
            "Sec-Fetch-Site": "same-origin",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36",
            "identity": self.ctx.call('indentify'),
            "sec-ch-ua": "\"Not)A;Brand\";v=\"99\", \"Google Chrome\";v=\"127\", \"Chromium\";v=\"127\"",
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": "\"Windows\""
        }
        request.cookies = {
            # "insert_cookie": "24796211"
        }
        tunnel = "n527.kdltps.com:15818"
        username = "t12360058344314"
        password = "cchhn0g9"
        request.proxies = {
            "http": f"http://{username}:{password}@{tunnel}/",
            "https": f"http://{username}:{password}@{tunnel}/"
        }
        return request

    def parse(self, request, response):
        count = 1
        data = self.ctx.call('decrypt', response.json.get('data'))
        data = json.loads(data)
        records = data['records']
        for record in records:
            # {'inscomname': '农银人寿保险股份有限公司', 'prodname': '农银人寿附加金穗团体公共保额医疗保险', 'prodstaticpagename': '8523ef14-d692-46d1-972f-0e70c4b2eec2', 'saledate': '2024-08-16', 'salestatus': '在售', 'startsaledate': 1723737600000, 'termsno': '8523ef14-d692-46d1-972f-0e70c4b2eec2', 'termsstaticpagename': '8523ef14-d692-46d1-972f-0e70c4b2eec2'}
            name = record['inscomname']
            prod_name = record['prodname']
            termsno = record['termsno']
            url = f'https://tiaokuan.iachina.cn/sinopipi/salemaintaapply/getHtmlData?termsno={termsno}'
            yield feapder.Request(url, method="GET", callback=self.parse_info,
                                  name=prod_name, count=count)
            count += 1
            # if count == 4:
            #     break


    def parse_info(self, request, response):
        zero = "0" * (4 - len(str(request.count)))
        file_name = zero + str(request.count) + request.name
        os.makedirs(f'./{self.top_file}/{file_name}', exist_ok=True)
        os.makedirs(f'./{self.top_file}/{file_name}/费率表', exist_ok=True)
        os.makedirs(f'./{self.top_file}/{file_name}/条款', exist_ok=True)

        # info {"prodVo":{"base":[{"name":"公司名称：","value":"农银人寿保险股份有限公司"},{"name":"产品名称：","value":"农银人寿附加金穗团体公共保额医疗保险"},{"name":"产品类别：","value":"健康保险-医疗保险-费用补偿型医疗保险-其他"},{"name":"设计类型：","value":"传统型产品-普通型"},{"name":"产品特殊属性：","value":"无"},{"name":"承保方式：","value":"团体"},{"name":"保险期间类型：","value":"短期（一年及一年以下）"},{"name":"产品交费方式：","value":"一次性交费"},{"name":"产品条款文字编码：","value":"农银人寿[2024]医疗保险022号"},{"name":"产品销售状态：","value":"在售"},{"name":"停止销售日期：","value":""},{"name":"客户服务电话：","value":"95581"}]},"termsVo":{"base":[{"name":"公司名称：","value":"农银人寿保险股份有限公司"},{"name":"产品名称：","value":"农银人寿附加金穗团体公共保额医疗保险"},{"name":"产品类别：","value":"健康保险-医疗保险-费用补偿型医疗保险-其他"},{"name":"设计类型：","value":"传统型产品-普通型"},{"name":"产品销售状态：","value":"在售"}],"file":[{"name":"费率文件：","type":"04","value":"26d12fad-9b93-4a4f-9923-47258a1ee2dc_RATE.PDF"}],"phone":[{"name":"客户服务电话：","value":"95581"}],"termsfilename":"3898180a-9f9b-4231-96c1-332502dfdbd8_TERMS.PDF","termsfiletype":"03","tip":""}}
        info = self.ctx.call('decrypt', response.json.get('data'))
        base = json.loads(info)['prodVo']['base']
        wb = Workbook()
        ws = wb.active
        for b in base:
            name, value = b.values()
            ws.cell(row=base.index(b) + 1, column=1, value=name)
            ws.cell(row=base.index(b) + 1, column=2, value=value)
        wb.save(f"./{self.top_file}/{file_name}/{request.name + '.xlsx'}")

        if json.loads(info)["termsVo"]["file"]:
            feilv_type = json.loads(info)["termsVo"]["file"][0]["type"]
            feilv_path = json.loads(info)["termsVo"]["file"][0]["value"]
            # {"fileType":"04","filePath":"26d12fad-9b93-4a4f-9923-47258a1ee2dc_RATE.PDF","pageSize":1,"fileName":"9a0582de7553acaea193e2687e10e79f","fleBasePath":"/cncfds/folder"}
            file_name_1 = '{"fileType":"' + feilv_type + '","filePath":"' + feilv_path + '","pageSize":1,"fileName":"9a0582de7553acaea193e2687e10e79f","fleBasePath":"/cncfds/folder"}'
            data_1 = {"filename": self.ctx.call('encrypt', file_name_1)}
            data_1 = json.dumps(data_1, separators=(',', ':'))
            url = "https://tiaokuan.iachina.cn/sinopipi/pdf/getPdfPage"
            yield feapder.Request(url, data=data_1, method="POST", callback=self.parse_feilv_pdf,
                                  feilv_type=feilv_type, feilv_path=feilv_path,
                                  name=request.name, file_name=file_name)

            tiaokuan_type = json.loads(info)["termsVo"]["termsfiletype"]
            tiaokuan_path = json.loads(info)["termsVo"]["termsfilename"]
            file_name_2 = '{"fileType":"' + tiaokuan_type + '","filePath":"' + tiaokuan_path + '","pageSize":1,"fileName":"9a0582de7553acaea193e2687e10e79f","fleBasePath":"/cncfds/folder"}'
            data_2 = {"filename": self.ctx.call('encrypt', file_name_2)}
            data_2 = json.dumps(data_2, separators=(',', ':'))
            url = "https://tiaokuan.iachina.cn/sinopipi/pdf/getPdfPage"
            yield feapder.Request(url, method="POST", data=data_2, callback=self.parse_tiaokuan_pdf,
                                  tiaokuan_type=tiaokuan_type, tiaokuan_path=tiaokuan_path,
                                  name=request.name, file_name=file_name)

            # 获取其他文件
            file = json.loads(info)["termsVo"]["file"]
            if len(file) > 1:
                for file in file[1:]:
                    name = file['name'].replace('：', '')
                    os.makedirs(f'./{self.top_file}/{file_name}/{name}', exist_ok=True)
                    type = file['type']
                    path = file['value']
                    file_name_3 = '{"fileType":"' + type + '","filePath":"' + path + '","pageSize":1,"fileName":"9a0582de7553acaea193e2687e10e79f","fleBasePath":"/cncfds/folder"}'
                    data_3 = {"filename": self.ctx.call('encrypt', file_name_3)}
                    data_3 = json.dumps(data_3, separators=(',', ':'))
                    url = "https://tiaokuan.iachina.cn/sinopipi/pdf/getPdfPage"
                    yield feapder.Request(url, method="POST", data=data_3, callback=self.parse_other_pdf,
                                          type=type, path=path, name=request.name, file_name=file_name, sub_name=name)

    def parse_other_pdf(self, request, response):
        info = response.json
        pdf = info['pdf']
        total_page = info['totalPage']
        path = f'./{self.top_file}/{request.file_name}/{request.sub_name}/{request.sub_name}{"0" * (len(str(total_page))-1)}1.gif'
        self.save_pdf(pdf, path)

        if total_page > 1:
            for i in range(2, total_page + 1):
                file_name = {
                    "fileType": f"{request.type}",
                    "filePath": f"{request.path}",
                    "pageSize": i,
                    "fileName": "9a0582de7553acaea193e2687e10e79f",
                    "fleBasePath": "/cncfds/folder"
                }
                data = {"filename": self.ctx.call('encrypt', str(file_name))}
                data = json.dumps(data, separators=(',', ':'))
                url = "https://tiaokuan.iachina.cn/sinopipi/pdf/getPdfPage"
                yield feapder.Request(url, data=data, method="POST", callback=self.parse_other_pdf_2,type=request.type,
                                      path=request.path, page=i, name=request.name, file_name=request.file_name, sub_name=request.sub_name)

    def parse_other_pdf_2(self, request, response):
        info = response.json
        pdf = info['pdf']
        total_page = info['totalPage']
        zero = len(str(total_page)) - len(str(request.page))
        path = f'./{self.top_file}/{request.file_name}/{request.sub_name}/{request.sub_name}{"0" * zero}{request.page}.gif'
        self.save_pdf(pdf, path)
        if total_page > 1:
            for i in range(2, total_page+1):
                file_name = {
                    "fileType": f"{request.type}",
                    "filePath": f"{request.path}",
                    "pageSize": i,
                    "fileName": "9a0582de7553acaea193e2687e10e79f",
                    "fleBasePath": "/cncfds/folder"
                }
                data = {"filename": self.ctx.call('encrypt', str(file_name))}
                data = json.dumps(data, separators=(',', ':'))
                url = "https://tiaokuan.iachina.cn/sinopipi/pdf/getPdfPage"
                yield feapder.Request(url, data=data, method="POST", callback=self.parse_other_pdf_3,
                                      page=i, name=request.name, file_name=request.file_name, sub_name=request.sub_name)

    def parse_other_pdf_3(self, request, response):
        info = response.json
        pdf = info['pdf']
        total_page = info['totalPage']
        zero = len(str(total_page)) - len(str(request.page))
        path = f'./{self.top_file}/{request.file_name}/{request.sub_name}/{request.sub_name}{"0" * zero}{request.page}.gif'
        self.save_pdf(pdf, path)


    def save_pdf(self, data, path):
        with open(f'{path}', 'wb') as f:
            f.write(base64.b64decode(self.ctx.call('decrypt', data)))

    def parse_feilv_pdf(self, request, response):
        info = response.json
        pdf = info['pdf']
        total_page = info['totalPage']
        path = f'./{self.top_file}/{request.file_name}/费率表/{request.name}_费率表{"0" * (len(str(total_page)) - 1)}1.gif'
        self.save_pdf(pdf, path)
        if total_page > 1:
            for i in range(2, total_page+1):
                file_name = {
                    "fileType": f"{request.feilv_type}",
                    "filePath": f"{request.feilv_path}",
                    "pageSize": i,
                    "fileName": "9a0582de7553acaea193e2687e10e79f",
                    "fleBasePath": "/cncfds/folder"
                }
                data = {"filename": self.ctx.call('encrypt', str(file_name))}
                data = json.dumps(data, separators=(',', ':'))
                url = "https://tiaokuan.iachina.cn/sinopipi/pdf/getPdfPage"
                yield feapder.Request(url, data=data, method="POST", callback=self.parse_feilv_pdf_2,
                                      page=i, name=request.name, file_name=request.file_name)

    def parse_feilv_pdf_2(self, request, response):
        info = response.json
        pdf = info['pdf']
        total_page = info['totalPage']
        zero = len(str(total_page)) - len(str(request.page))
        path = f'./{self.top_file}/{request.file_name}/费率表/{request.name}_费率表{"0" * zero}{request.page}.gif'
        self.save_pdf(pdf, path)

    def parse_tiaokuan_pdf(self, request, response):
        data = response.json
        pdf = data['pdf']
        total_page = data['totalPage']
        path = f'./{self.top_file}/{request.file_name}/条款/{request.name}_条款{"0" * (len(str(total_page)) - 1)}1.gif'
        self.save_pdf(pdf, path)
        if total_page > 1:
            for i in range(2, total_page+1):
                file_name = {
                    "fileType": f"{request.tiaokuan_type}",
                    "filePath": f"{request.tiaokuan_path}",
                    "pageSize": i,
                    "fileName": "9a0582de7553acaea193e2687e10e79f",
                    "fleBasePath": "/cncfds/folder"
                }
                data = {"filename": self.ctx.call('encrypt', str(file_name))}
                data = json.dumps(data, separators=(',', ':'))
                url = "https://tiaokuan.iachina.cn/sinopipi/pdf/getPdfPage"
                yield feapder.Request(url, data=data, method="POST", callback=self.parse_tiaokuan_pdf_2,
                                      page=i, name=request.name, file_name=request.file_name)

    def parse_tiaokuan_pdf_2(self, request, response):
        info = response.json
        pdf = info['pdf']
        total_page = info['totalPage']
        zero = len(str(total_page)) - len(str(request.page))
        path = f'./{self.top_file}/{request.file_name}/条款/{request.name}_条款{"0" * zero}{request.page}.gif'
        self.save_pdf(pdf, path)


if __name__ == "__main__":
    AirSpiderDemo(thread_count=5).start()
